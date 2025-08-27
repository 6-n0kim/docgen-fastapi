from openpyxl import Workbook
from ai.product_requirement_document.prd_generator import Requirement_Document
from ai.functional_specification_document.fsd_generator import FunctionalSpecification
def create_prd(requirement_document_data : Requirement_Document):
  wb = Workbook()
  ws = wb.active

  ws.title = "요구사항정의서"
  column = ['ID', '시스템', '기능', '요구사항', '설명']
  ws.append(column)

  datas = requirement_document_data.data
  for data in datas:
    for detail in data.details:
      row = ['-','-',data.name, detail.name, detail.description]
      ws.append(row)

  merge_column_cell(ws,3)

  wb.save("요구사항정의서.xlsx")
  wb.close()

def create_fsd(functional_specification_data : FunctionalSpecification):
  wb = Workbook()
  ws = wb.active

  ws.title = "기능명세서"
  column = ['ID', '시스템', '메뉴', '기능', '설명']
  ws.append(column)

  datas = functional_specification_data.data
  for data in datas:
    for detail in data.details:
      row = ['-','-',data.name, detail.name, detail.description]
      ws.append(row)

  merge_column_cell(ws,3)

  wb.save("기능명세서.xlsx")
  wb.close()

def merge_column_cell(ws, col_idx,start_row = 2):
  max_row = ws.max_row

  merge_start = start_row
  prev_value = ws.cell(row=start_row, column=col_idx).value

  for row in range(start_row + 1, max_row + 1):
      cell_value = ws.cell(row=row, column=col_idx).value
      if cell_value != prev_value:
          # 이전까지 동일한 값 병합
          if row - 1 > merge_start:
              ws.merge_cells(start_row=merge_start, start_column=col_idx,end_row=row - 1, end_column=col_idx)
          merge_start = row
          prev_value = cell_value
  # 마지막 값 처리
  if max_row > merge_start:
      ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=max_row, end_column=col_idx)